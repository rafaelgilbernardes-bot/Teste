"""POST /api/relatorios/excel

Gera o relatório Excel no formato atual do escritório (briefing § 11).

Estrutura:
    Cliente: [nome]           Referência: [Mês/Ano]
    Valor da hora: R$ [X]     Valor total: R$ [calculado]

    | Data | Responsável | Projeto | Descrição | Duração | Valor |
    ...

    Valor dos serviços prestados:    [total horas] [total R$]
    Valor total para faturamento:    -             R$ [total]
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db import get_db

router = APIRouter(tags=["relatorios"])

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BRL = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


@router.get("/relatorios/excel")
def gerar_relatorio_excel(
    contrato_id: str = Query(...),
    mes: str = Query(..., description="YYYY-MM"),
):
    db = get_db()

    # Buscar contrato e cliente
    c_resp = (
        db.table("contratos")
        .select("*, clientes(nome, contato_nome)")
        .eq("id", contrato_id)
        .limit(1)
        .execute()
    )
    if not c_resp.data:
        return {"error": "Contrato não encontrado"}
    contrato = c_resp.data[0]
    cliente_nome = (contrato.get("clientes") or {}).get("nome", "?")

    # Buscar entries do mês
    entries = (
        db.table("time_entries")
        .select("data, duracao_minutos, tarefa_nome, descricao, providencia, "
                "colaboradores(nome)")
        .eq("contrato_id", contrato_id)
        .eq("mes_referencia", mes)
        .eq("alerta_sem_entry", False)
        .not_.is_("produto", "null")
        .order("data")
        .execute()
        .data
    )

    # Cálculos
    modelo = contrato["modelo"]
    valor_hora = contrato.get("valor_hora") or 0
    total_min = sum(e["duracao_minutos"] for e in entries)
    total_h = total_min / 60

    if modelo == "hora":
        valor_total = total_h * valor_hora
    elif modelo == "laas":
        valor_total = contrato.get("valor_fixo_mensal") or 0
    else:
        valor_total = contrato.get("valor_escopo") or 0

    # Formatar mês
    mes_dt = datetime.strptime(mes, "%Y-%m")
    mes_label = mes_dt.strftime("%B/%Y").capitalize()

    # Montar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Relatório {mes_label}"

    # Cabeçalho do relatório
    ws.append([f"Cliente: {cliente_nome}", "", "", f"Referência: {mes_label}"])
    ws.append([f"Modelo: {modelo.replace('_', ' ').title()}", "", "",
               f"Valor da hora: {_BRL(valor_hora)}" if modelo == "hora" else ""])
    ws.append([])

    # Cabeçalho da tabela
    headers = ["Data", "Responsável", "Projeto", "Descrição", "Duração", "Valor"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(header_row, col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Linhas de atividade
    for e in entries:
        dur_min = e["duracao_minutos"]
        dur_h = dur_min / 60
        h = int(dur_h)
        m = int((dur_h - h) * 60)
        dur_label = f"{h}h{m:02d}m"

        if modelo == "hora":
            valor_linha = dur_h * valor_hora
        else:
            valor_linha = None

        colab_nome = (e.get("colaboradores") or {}).get("nome", "")
        # Descrição: preferir providencia + tarefa_nome (substantivos)
        descricao = e.get("descricao") or ""
        if not descricao and e.get("providencia"):
            descricao = f"{e['providencia']} — {e.get('tarefa_nome', '')}"

        ws.append([
            e["data"],
            colab_nome,
            e.get("tarefa_nome", ""),
            descricao,
            dur_label,
            _BRL(valor_linha) if valor_linha is not None else "",
        ])

    # Rodapé
    ws.append([])
    h_int = int(total_h)
    h_min = int((total_h - h_int) * 60)
    ws.append(["Valor dos serviços prestados:", "", "", "",
               f"{h_int}h{h_min:02d}m", _BRL(valor_total)])
    ws.append(["Reembolso de despesas:", "", "", "", "-", _BRL(0)])
    ws.append(["Valor total para faturamento:", "", "", "", "-", _BRL(valor_total)])

    # Ajustar larguras
    col_widths = [12, 25, 35, 50, 10, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Salvar em buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"relatorio_{cliente_nome.replace(' ', '_')}_{mes}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
